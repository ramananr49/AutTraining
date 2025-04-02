import openpyxl

book = openpyxl.load_workbook("C:\\SeleniumTraining\\ExcelDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=1)
print(cell.value)
#Print Max row and column count

wb = openpyxl.load_workbook("C:\\SeleniumTraining\\ExcelDemo.xlsx")
# ws = wb.get_sheet_names()
ws = wb['Sheet1']

print(ws.max_row)
print(ws.max_column)

for i in range(1, ws.max_row+1):
    for j in range(1, ws.max_column):
        print(ws.cell(row=i, column=j).value)

